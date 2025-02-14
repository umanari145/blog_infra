import openpyxl
from jinja2 import Environment, FileSystemLoader

# Excelファイルの読み込み
workbook = openpyxl.load_workbook('api_paths.xlsx')
resource_sheet = workbook['resource']
method_sheet = workbook['method']
api_name = "blog_api2"
description="skill-up-engineering.comのblog"

# データの読み込み
resources = {}
methods = []

for row in resource_sheet.iter_rows(min_row=2, values_only=True):
    path_parts = row[1].split('/')
    parent_path = '/'.join(path_parts[:-1])
    resource = {
        'id': row[0],
        'path': path_parts[-1],
        'full_path': row[1],
        'parent_path': parent_path if parent_path else None
    }
    resources[row[1]] = resource

for row in method_sheet.iter_rows(min_row=2, values_only=True):
    methods.append({'id': row[0], 'resource_id': row[1], 'method': row[2], 'lambda': row[3]})

# Jinjaテンプレートの設定
env = Environment(loader=FileSystemLoader('.'))
template = env.get_template('template2.tf.j2')

# Terraformソースの生成
output = template.render(
    api_name=api_name,
    description=description,    
    resources=resources,
    methods=methods
)

# 出力ファイルの書き込み
with open('apigateway2.tf', 'w') as f:
    f.write(output)

print("Terraformソースが生成されました。")
